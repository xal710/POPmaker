import re
import zipfile
from pathlib import Path

import openpyxl

XLSM = Path(r"c:\Users\youse\Downloads\買取価格比較\買取価格比較_ver1.2.xlsm")

with zipfile.ZipFile(XLSM) as z:
    xml = z.read("xl/connections.xml").decode("utf-8")
    urls = sorted(set(re.findall(r"https://[^<\"']+", xml)))
    print("URLs in connections:")
    for u in urls:
        print(" ", u.replace("&amp;", "&"))

wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=True)

HARERUYA_SPEC_RE = re.compile(r"\([^)]*\)\{[^}]*\}")


def normalize_hareruya(name: str) -> str:
    return HARERUYA_SPEC_RE.sub("", name)


def normalize_cardrush(name: str, pack: str | None, model: str | None) -> str:
    card_name = re.sub(r"([^\s])\(", r"\1 (", str(name))
    model = str(model or "").strip()
    pack = str(pack or "").strip()
    body = f"{card_name}〈{model}〉"
    if pack and pack != "その他":
        return f"{body}[{pack}]"
    return body


raw_h = [(r[0], r[1]) for r in wb["20260607_晴れる屋2"].iter_rows(min_row=2, values_only=True) if r[0]]
chg_h = {r[0]: r[1] for r in wb["20260607_晴れる屋2_変更"].iter_rows(min_row=2, values_only=True) if r[0]}

h_ok = h_miss = 0
h_miss_samples = []
for name, price in raw_h:
    pred = normalize_hareruya(name)
    actual = chg_h.get(pred)
    if actual is not None and actual == price:
        h_ok += 1
    else:
        h_miss += 1
        if len(h_miss_samples) < 15:
            h_miss_samples.append((name, pred, actual))

print(f"\nHareruya normalize: ok={h_ok} miss={h_miss} total={len(raw_h)}")
for s in h_miss_samples:
    print(" ", s)

raw_c = [
    r for r in wb["20260607_カードラッシュ"].iter_rows(min_row=2, values_only=True) if r[0]
]
chg_c = {r[0]: r[1] for r in wb["20260607_カードラッシュ_変更"].iter_rows(min_row=2, values_only=True) if r[0]}

c_ok = c_miss = 0
c_miss_samples = []
for row in raw_c:
    name, pack, _rarity, model, price = row[0], row[1], row[2], row[3], row[4]
    pred = normalize_cardrush(name, pack, model)
    actual = chg_c.get(pred)
    if actual is not None and actual == price:
        c_ok += 1
    else:
        c_miss += 1
        if len(c_miss_samples) < 15:
            c_miss_samples.append((row, pred, actual))

print(f"\nCardrush normalize: ok={c_ok} miss={c_miss} total={len(raw_c)}")
for s in c_miss_samples:
    print(" ", s)

wb.close()
